import os
import re
import time

from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from win10toast import ToastNotifier
from openpyxl import load_workbook


class App:
    def __init__(self):
        
        load_dotenv()

        self.notifier = ToastNotifier()

        self.wb = load_workbook("RJ.xlsx")
        self.sheet = self.wb.active

        os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

        #   Configurar Navegador
        options = Options()        
        options.add_argument("--start-maximized")
        options.add_argument("--log-level=3")
        options.add_argument("--disable-logging")
        options.add_argument("--disable-dev-shm-usage")
        
        self.navegador = webdriver.Chrome()
        self.navegador.get(os.getenv("LINK"))
        self.navegador.maximize_window()

        self.run()
        
    def navegar(self):
        
        time.sleep(5)
        self.navegador.refresh()

        #   NAVEGA PELO MENU DO SITE
        try:
            botao_menu = WebDriverWait(self.navegador, 15).until(
                EC.presence_of_element_located((By.CLASS_NAME, "botao-menu"))
            )
            botao_menu.click()
            print("\nNavegando pelo site...")
            time.sleep(1)  # Espera um pouco para garantir que o menu esteja aberto

            processo = self.navegador.find_element("partial link text", "Processo")
            processo.click()
            time.sleep(.5)

            acoes = self.navegador.find_elements(By.XPATH, "//*[contains(text(), ' Outras ações ')]")	
            acoes[0].click()
            time.sleep(.5)

            soliciar_hab = self.navegador.find_elements(By.XPATH, "//*[contains(text(), ' Solicitar habilitação ')]")
            soliciar_hab[0].click()
            time.sleep(.5)
            
        except Exception as e:

            self.notifier.show_toast("ERRO", "Erro ao navegar pelo site", duration=3)
            time.sleep(5)
            raise Exception("Erro ao navegar pelo site") from e

    def ponteiro(self):
        for row in self.sheet.iter_rows(min_row=2, max_col=1):
            cell_a = row[0]
            num_peticao = str(cell_a.value).strip()
            self.linha = cell_a.row

            yield num_peticao

    def atualizar_xlsx(self, coluna, valor):
        self.sheet[f"{coluna}{self.linha}"] = valor

    def getNumPeticao(self, num_peticao):
        
        try:
            peticao_str = str(num_peticao)

            self.notifier.show_toast("Início", f"Buscando por: \n{peticao_str}", duration=4)
            print("====================================================================")
            print(f"Buscando por: {num_peticao}\n\n")

            peticao_str = re.sub(r'[^0-9]','', peticao_str)

            if peticao_str.isdigit() != True:
                print("\n\nPetição não é valida, indo para a proxima.\n\n")
                return False

            time.sleep(1)
            self.navegador.find_element(By.ID, "fPP:numeroProcesso:numeroSequencial").send_keys(peticao_str[0:7])
            self.navegador.find_element(By.ID, "fPP:numeroProcesso:numeroDigitoVerificador").send_keys(peticao_str[7:9])
            self.navegador.find_element(By.ID, "fPP:numeroProcesso:Ano").send_keys(peticao_str[9:13])
            self.navegador.find_element(By.ID, "fPP:numeroProcesso:NumeroOrgaoJustica").send_keys(peticao_str[-4:])
            
            # Pesquisar Processo
            time.sleep(1)
            self.navegador.find_element(By.ID, "fPP:searchProcessos").click()

            time.sleep(1)
            btn_link = WebDriverWait(self.navegador, 10).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div/div/div/div[2]/form/div[2]/div[2]/table/tbody/tr/td[2]/a'))
            )
            btn_link.click()

        except:
            
            self.notifier.show_toast("ERRO", "Erro ao pesquisar por Petição", duration=3)
            time.sleep(5)
            print("Erro ao pesquisar petição\n")
            raise      

    def getPolo(self):
        try:
            time.sleep(3)
            abas = self.navegador.window_handles
            self.navegador.switch_to.window(abas[1])

            if os.getenv('NOME_DO_POLO') in self.navegador.page_source:
                self.atualizar_xlsx("B", "ATIVO")
                print("\n✅ POLO ATIVO\n")
                return True
        
        except:
            print("❌ Não foi possível localizar o polo ativo")
            self.atualizar_xlsx("B", "INATIVO")
            self.navegador.switch_to.alert.dismiss()
            return False

    def getStatus(self):

        var = False

        if "arquivado" in self.navegador.page_source:
            self.atualizar_xlsx("C","ARQUIVADO")
            print("✅ Caso está ARQUIVADO")
            var = True
            
        if "baixado" in self.navegador.page_source:
            self.atualizar_xlsx("D","BAIXADO")
            print("✅ Caso está BAIXADO")
            var = True

        if "setença" in self.navegador.page_source:
            self.atualizar_xlsx("E","SENTENÇA")
            print("✅ Caso está SENTENCIADO")
            var = True

        if "suspenso" in self.navegador.page_source:
            self.atualizar_xlsx("F","SUSPENSO")
            print("✅ Caso está SUSPENSO")
            var = True

        if var == False:
            print("🟨 NENHUM STATUS ENCONTRADO!!")

    def fim(self):      
        abas = self.navegador.window_handles
        if len(abas) > 1:
            self.navegador.close()

        self.navegador.switch_to.window(abas[0])

        self.navegador.find_element(By.ID, "fPP:clearButtonProcessos").click()
        time.sleep(2)

    def run(self):
        self.navegar()

        for num_peticao in self.ponteiro():
            self.getNumPeticao(num_peticao)

            if self.getPolo() == True:
                self.getStatus()

            self.wb.save("RJ.xlsx")
            print("\n✅ Alteração Salva!!\n\n")
            self.fim()

try:
    print('incio')
    app = App()
    print('fim')
    
except Exception as e:
    print(f"\n\nERRO AO EXECUTAR AUTOMAÇÃO\n\n")
    print(e)
    
    time.sleep(5)